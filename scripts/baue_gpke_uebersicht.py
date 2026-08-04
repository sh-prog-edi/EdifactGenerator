# Erzeugt die prozessweise GPKE-Übersicht (Initial-/Antwortnachrichten) als XLSX.
# Quellen: GPKE Teil 1-4 (BNetzA, BK6-24-174 bzw. BK6-22-024, Reihenfolge der
# Use-Cases) + EDI@Energy "Anwendungsübersicht der Prüfidentifikatoren 3.3"
# (kons. 29.06.2026, BDEW-MAKO fileId 12258, Tabelle 1).
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

QUELLE = "regelwerk/Anwendungsuebersicht_3.3_LF_12258.xlsx"
ZIEL = "GPKE_Prozess-Nachrichten-Uebersicht.xlsx"

# Use-Case-Reihenfolge gemäß GPKE-Inhaltsverzeichnissen (BK6-24-174 / BK6-22-024)
ORDNUNG = {
    "GPKE Teil 2": [
        "Ermittlung der MaLo-ID", "Kündigung", "Lieferbeginn", "Neuanlage",
        "Beginn der Ersatz-/Grundversorgung",
        "Fall 1: LF-Zuordnung", "Fall 2: LF-Zuordnung", "Fall 3: LF-Zuordnung",
        "Fall 4: LF-Zuordnung",
        "Lieferende von LF an NB", "Lieferende von NB an LF",
        "Abrechnungsdaten Netznutzungsabrechnung",
        "Abrechnungsdaten Bilanzkreisabrechnung",
        "Bestellung einer Änderung von Abrechnungsdaten von LF an NB",
        "Bestellung einer Änderung von Abrechnungsdaten zur Bilanzkreisabrechnu",
        "Übermittlung der bisher gemessenen",
        "Übermittlung des Lieferscheins",
        "Netznutzungsabrechnung",
        "Übermittlung Preisblatt NB an LF",
        "Abrechnung einer sonstigen Leistung",
        "Unterbrechung der Anschlussnutzung (Sperren)",
        "Wiederherstellung der Anschlussnutzung (Entsperren)",
        "Stornieren der Unterbrechung",
        "Wiederherstellung der Anschlussnutzung bei Lieferbeginn",
    ],
    "GPKE Teil 3": [
        "Übermittlung der Übersicht der Definitionen des NB",
        "Übermittlung der Übersicht der Definitionen des LF",
        "Übermittlung einer Definition des NB",
        "Übermittlung einer Definition des LF",
        "Reklamation der Übersicht der Definitionen des NB vom LF",
        "Reklamation der Übersicht der Definitionen des NB vom MSB",
        "Reklamation der Übersicht der Definitionen des LF vom NB",
        "Reklamation der Übersicht der Definitionen des LF vom MSB",
        "Reklamation einer Definition des NB vom LF",
        "Reklamation einer Definition des NB vom MSB",
        "Reklamation einer Definition des LF vom NB",
        "Reklamation einer Definition des LF vom MSB",
        "Bestellung einer Konfiguration vom LF an NB",
        "Bestellung einer Konfiguration vom NB an MSB",
        "Bestellung einer Konfiguration vom LF an MSB",
        "Reklamation einer Konfiguration vom NB an MSB",
        "Reklamation einer Konfiguration vom LF an MSB",
        "Reklamation einer Konfiguration vom MSB",
        "Bestellung Beendigung einer Konfiguration vom NB an MSB",
        "Bestellung Beendigung einer Konfiguration vom LF an MSB",
        "Bestellung Beendigung einer Konfiguration vom weiteren MSB",
        "Beendigung einer Konfiguration vom MSB",
        "Einrichtung der Konfigurationen",
        "Steuerbefehl vom NB an MSB",
        "Steuerbefehl vom LF an MSB (UF B016)",
        "Steuerbefehl vom LF an MSB (UF B024)",
        "Übermittlung Preisblatt A des MSB",
        "Abrechnung Leistungen des Preisblatts A des MSB zwischen MSB und NB",
        "Abrechnung Leistungen des Preisblatts A des MSB zwischen MSB und LF",
    ],
    "GPKE Teil 4": [
        "Stammdatenänderung vom NB", "Stammdatenänderung vom LF",
        "Stammdatenänderung vom MSB", "Stammdatenänderung vom ÜNB",
        "Bestellung zur Stammdatenänderung an NB",
        "Bestellung zur Stammdatenänderung an LF",
        "Bestellung zur Stammdatenänderung an MSB",
        "Bestellung zur Stammdatenänderung an ÜNB",
        "Stammdaten zur Bilanzkreistreue",
        "Geschäftsdatenanfrage",
        "Übermittlung von Informationen",
        "Stornierung",   # Kap. 5 (Zeilen ohne SD-Bezeichnung)
    ],
}


def norm(s):
    return re.sub(r"\s+", " ", str(s or "")).strip()


def nachrichtentyp(ahb, api):
    a = norm(ahb)
    if a in ("--", "") and api:
        return "API"
    m = re.match(r"([A-Z]{6}|CONTRL|APERAK)", a)
    return m.group(1) if m else (a.split(" ")[0] if a else "--")


wb = openpyxl.load_workbook(QUELLE, read_only=True)
ws = wb["Prüf-ID Prozessschritt"]
rows = list(ws.iter_rows(values_only=True))
kopf = [norm(x) for x in rows[0]]

zeilen = []
for r in rows[1:]:
    pb = norm(r[5])
    if not pb.startswith("GPKE"):
        continue
    teil = pb[:11].strip()          # "GPKE Teil 2/3/4"
    sd = norm(r[7]) or ("Stornierung (Kap. 5)" if "Kap. 5" in norm(r[6]) else norm(r[6]))
    zeilen.append({
        "teil": teil, "sd": sd,
        "schritt": norm(r[8]) or "–",
        "aktion": norm(r[9]) if norm(r[9]) not in ("--", "") else norm(r[2]),
        "von": norm(r[10]), "an": norm(r[11]),
        "typ": nachrichtentyp(r[1], norm(r[19]) not in ("", "--")),
        "ahb": norm(r[1]), "fall": norm(r[2]),
        "pruefi": norm(r[3]), "reaktion": norm(r[4]),
        "weg": norm(r[18]), "api": norm(r[19]),
    })

# Prozess-Reihenfolge je Teil bestimmen
def sd_rang(teil, sd):
    liste = ORDNUNG.get(teil, [])
    n = norm(sd).lower()
    for i, muster in enumerate(liste):
        if n.startswith(muster.lower()[:min(len(muster), 38)]):
            return i
    return 900

def schritt_key(z):
    m = re.match(r"(\d+)", z["schritt"])
    return (int(m.group(1)) if m else 99, z["pruefi"])

# ---- Arbeitsmappe ------------------------------------------------------------
out = openpyxl.Workbook()
ARIAL = "Arial"
KOPF_FILL = PatternFill("solid", fgColor="1A4D8F")
PROZ_FILL = PatternFill("solid", fgColor="D9E2F0")
INIT_FILL = PatternFill("solid", fgColor="E6F5EC")
duenn = Side(style="thin", color="C0C8D4")
RAHMEN = Border(left=duenn, right=duenn, top=duenn, bottom=duenn)

SPALTEN = ["Prozess (Use-Case / SD)", "Schritt", "Aktion (Nachricht im SD)",
           "von", "an", "Nachrichtentyp", "Prüf-ID / API",
           "Rolle im Prozess", "Reaktion auf Prüf-ID",
           "Anwendungsfall lt. AHB", "AHB", "Übertragungsweg"]
BREITEN = [34, 7, 34, 12, 12, 14, 12, 22, 16, 34, 20, 12]

uebersicht = []   # (teil, sd, initial, antworten, typen)

for teil in ("GPKE Teil 2", "GPKE Teil 3", "GPKE Teil 4"):
    blatt = out.create_sheet(teil)
    for c, (name, breite) in enumerate(zip(SPALTEN, BREITEN), 1):
        z = blatt.cell(1, c, name)
        z.font = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
        z.fill = KOPF_FILL
        z.alignment = Alignment(wrap_text=True, vertical="center")
        blatt.column_dimensions[openpyxl.utils.get_column_letter(c)].width = breite
    blatt.freeze_panes = "A2"
    r = 2
    daten = [z for z in zeilen if z["teil"] == teil]
    sds = sorted({z["sd"] for z in daten}, key=lambda s: (sd_rang(teil, s), s))
    for nr, sd in enumerate(sds, 1):
        block = sorted([z for z in daten if z["sd"] == sd], key=schritt_key)
        # Prozess-Kopfzeile
        z = blatt.cell(r, 1, f"{nr}. {sd}")
        z.font = Font(name=ARIAL, bold=True, size=10)
        z.fill = PROZ_FILL
        for c in range(1, len(SPALTEN) + 1):
            blatt.cell(r, c).fill = PROZ_FILL
            blatt.cell(r, c).border = RAHMEN
        r += 1
        initial, antworten, typen = [], [], set()
        for z0 in block:
            # Legende Anwendungsübersicht: konkrete Prüf-ID(s) = Antwort auf diese;
            # "x" = referenzierter Auslöser (auf ihn wird geantwortet); "--" = keine
            # 1:n-Disambiguierung nötig -> Antwort am Aktionstext erkennbar
            reakt = z0["reaktion"].replace("\n", ", ")
            ist_antwort_expl = reakt not in ("--", "", "x")
            ist_antwort_heur = bool(re.match(
                r"^(Antwort|Bestätigung|Ablehnung|Zustimmung|Abweisung|Rückmeldung)\b",
                z0["aktion"], re.I)) and z0["schritt"] != "1"
            if z0["schritt"] == "–" and not ist_antwort_expl:
                ist_antwort_heur = bool(re.match(
                    r"^(Antwort|Bestätigung|Ablehnung|Zustimmung|Abweisung|Rückmeldung)\b",
                    z0["aktion"], re.I))
            ist_antwort = ist_antwort_expl or ist_antwort_heur
            ist_initial = (z0["schritt"] == "1" or (z0["schritt"] == "–" and not ist_antwort)) and not ist_antwort
            rolle = ("Initialnachricht" if ist_initial
                     else (f"Antwort auf {reakt}" if ist_antwort_expl
                           else ("Antwort (auf vorherigen Schritt)" if ist_antwort_heur
                                 else "Folgemeldung")))
            z0["reaktion"] = reakt if ist_antwort_expl else ("--" if not ist_antwort_heur else "vorheriger Schritt")
            kennung = z0["pruefi"] if z0["pruefi"] not in ("--", "") else (z0["api"] if z0["api"] not in ("--", "") else "–")
            werte = [None, z0["schritt"], z0["aktion"], z0["von"], z0["an"], z0["typ"],
                     kennung, rolle, z0["reaktion"] if ist_antwort else "–",
                     z0["fall"], z0["ahb"], z0["weg"]]
            for c, w in enumerate(werte, 1):
                zz = blatt.cell(r, c, w)
                zz.font = Font(name=ARIAL, size=10)
                zz.border = RAHMEN
                zz.alignment = Alignment(wrap_text=(c in (3, 10)), vertical="top")
            ausgepraegt = kennung != "–" and z0["typ"] not in ("--", "")
            if ist_initial:
                for c in range(1, len(SPALTEN) + 1):
                    blatt.cell(r, c).fill = INIT_FILL
                if ausgepraegt: initial.append(f"{z0['typ']} {kennung}")
            elif ist_antwort and ausgepraegt:
                antworten.append(f"{z0['typ']} {kennung}")
            if ausgepraegt: typen.add(z0["typ"])
            r += 1
        uebersicht.append((teil, f"{nr}. {sd}", initial, antworten, typen, len(block)))
    print(f"{teil}: {len(sds)} Prozesse, {len(daten)} Nachrichten/Schritte")

# ---- Übersichtsblatt ---------------------------------------------------------
ub = out.active
ub.title = "Übersicht"
UEB = ["GPKE-Teil", "Prozess (in Dokumentreihenfolge)", "Initialnachricht(en)",
       "direkte Antwortnachrichten", "beteiligte Nachrichtentypen", "Schritte gesamt"]
for c, (name, breite) in enumerate(zip(UEB, [12, 44, 30, 40, 26, 10]), 1):
    z = ub.cell(1, c, name)
    z.font = Font(name=ARIAL, bold=True, color="FFFFFF", size=10)
    z.fill = KOPF_FILL
    z.alignment = Alignment(wrap_text=True, vertical="center")
    ub.column_dimensions[openpyxl.utils.get_column_letter(c)].width = breite
ub.freeze_panes = "A2"
for r, (teil, sd, initial, antworten, typen, n) in enumerate(uebersicht, 2):
    werte = [teil, sd, "; ".join(dict.fromkeys(initial)) or "–",
             "; ".join(dict.fromkeys(antworten)) or "–",
             ", ".join(sorted(typen)), n]
    for c, w in enumerate(werte, 1):
        z = ub.cell(r, c, w)
        z.font = Font(name=ARIAL, size=10)
        z.border = RAHMEN
        z.alignment = Alignment(wrap_text=(c in (2, 3, 4, 5)), vertical="top")

# Quellenblatt
qb = out.create_sheet("Quellen")
qtexte = [
    "GPKE Teil 1-3, Lesefassungen BK6-24-174 (Beschluss 24.10.2024, gültig ab 06.06.2025), Bundesnetzagentur.",
    "GPKE Teil 4 (Fokus Stammdatenprozesse), Anlage 1d zu BK6-22-024, Bundesnetzagentur.",
    "EDI@Energy Anwendungsübersicht der Prüfidentifikatoren 3.3, konsolidierte Lesefassung Stand 29.06.2026 (BDEW-MAKO, Tabelle 1) - Zuordnung Prüf-ID/Prozessschritt, Spalte 'Reaktion auf Prüfidentifikator'.",
    "Reihenfolge der Prozesse gemäß Inhaltsverzeichnissen der GPKE-Teile; Formatstand 202604 (Version 3.3). Für 202610 liegt Version 4.0 vor.",
    "Rolle im Prozess: Initialnachricht = Prozessschritt 1 ohne Reaktionsbezug; Antwort = Zeile mit 'Reaktion auf Prüfidentifikator'; Folgemeldung = weitere Meldung ohne direkten Reaktionsbezug (z.B. Weiterleitungen, Info-Meldungen).",
]
qb.cell(1, 1, "Quellen und Lesehinweise").font = Font(name=ARIAL, bold=True, size=11)
for i, t in enumerate(qtexte, 3):
    z = qb.cell(i, 1, "• " + t)
    z.font = Font(name=ARIAL, size=10)
    z.alignment = Alignment(wrap_text=True)
qb.column_dimensions["A"].width = 130

out.save(ZIEL)
print("->", ZIEL)
